from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from datetime import date
from typing import Optional
import time
import io
import csv

from app.services.invoice_db import query_invoices, iter_rows, get_data, safe_float, get_review_queue

router = APIRouter(prefix="/stats", tags=["Stats"])

# ─── Basit in-memory cache (60 sn TTL) ───────────────────
_cache: dict = {}
_CACHE_TTL = 60


def _cache_key(*args) -> str:
    return str(args)


def _cached(key: str, fn):
    now = time.time()
    if key in _cache and now - _cache[key]["ts"] < _CACHE_TTL:
        return _cache[key]["data"]
    result = fn()
    _cache[key] = {"ts": now, "data": result}
    return result


def _invalidate():
    _cache.clear()


# ─── GET /stats/total ─────────────────────────────────────
@router.get("/total")
def total():
    def _calc():
        r = query_invoices(per_page=1)
        return {
            "count":       r["count"],
            "total_sum":   r["total_sum"],
            "vat_sum":     r["vat_sum"],
            "by_vendor":   r["by_vendor"],
            "by_category": r["by_category"],
        }
    return _cached("total", _calc)


# ─── GET /stats/summary  (kombine filtre + pagination) ────
@router.get("/summary")
def summary(
    start:      Optional[date]  = Query(None),
    end:        Optional[date]  = Query(None),
    vendor:     Optional[str]   = Query(None, max_length=100),
    category:   Optional[str]   = Query(None, max_length=50),
    payment:    Optional[str]   = Query(None, max_length=50),
    invoice_no: Optional[str]   = Query(None, max_length=100),
    min_amount: Optional[float] = Query(None, ge=0),
    max_amount: Optional[float] = Query(None, ge=0),
    page:       int             = Query(1, ge=1),
    per_page:   int             = Query(100, ge=1, le=500),
):
    r = query_invoices(
        start=str(start) if start else None,
        end=str(end) if end else None,
        vendor=vendor,
        category=category,
        payment=payment,
        invoice_no=invoice_no,
        min_amt=min_amount,
        max_amt=max_amount,
        page=page,
        per_page=per_page,
    )
    return r


# ─── GET /stats/by-date ───────────────────────────────────
@router.get("/by-date")
def by_date(
    start:    date = Query(...),
    end:      date = Query(...),
    page:     int  = Query(1, ge=1),
    per_page: int  = Query(100, ge=1, le=500),
):
    r = query_invoices(start=str(start), end=str(end), page=page, per_page=per_page)
    return {"start": str(start), "end": str(end), **r}


# ─── GET /stats/by-vendor ─────────────────────────────────
@router.get("/by-vendor")
def by_vendor(
    vendor:   str = Query(..., max_length=100),
    page:     int = Query(1, ge=1),
    per_page: int = Query(100, ge=1, le=500),
):
    r = query_invoices(vendor=vendor, page=page, per_page=per_page)
    return {"vendor": vendor, **r}


# ─── GET /stats/by-category ───────────────────────────────
@router.get("/by-category")
def by_category(
    category: str = Query(..., max_length=50),
    page:     int = Query(1, ge=1),
    per_page: int = Query(100, ge=1, le=500),
):
    r = query_invoices(category=category, page=page, per_page=per_page)
    return {"category": category, **r}


# ─── GET /stats/by-payment ────────────────────────────────
@router.get("/by-payment")
def by_payment(
    method:   str = Query(..., max_length=50),
    page:     int = Query(1, ge=1),
    per_page: int = Query(100, ge=1, le=500),
):
    r = query_invoices(payment=method, page=page, per_page=per_page)
    return {"payment_method": method, **r}


# ─── GET /stats/by-invoice-no ─────────────────────────────
@router.get("/by-invoice-no")
def by_invoice_no(invoice_no: str = Query(..., max_length=100)):
    r = query_invoices(invoice_no=invoice_no, per_page=200)
    return {"invoice_no": invoice_no, "count": r["count"], "invoices": r["invoices"]}


# ─── GET /stats/export/excel  ─────────────────────────────
# openpyxl write_only = streaming writer — RAM sabit (N→∞)
# Excel hard limit: 1.048.576 satır → aşıldığında CSV önerilir
_EXCEL_ROW_LIMIT = 1_048_575  # başlık satırı hariç

@router.get("/export/excel")
def export_excel(
    start:      Optional[date]  = Query(None),
    end:        Optional[date]  = Query(None),
    vendor:     Optional[str]   = Query(None, max_length=100),
    category:   Optional[str]   = Query(None, max_length=50),
    min_amount: Optional[float] = Query(None, ge=0),
    max_amount: Optional[float] = Query(None, ge=0),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "openpyxl kurulu değil"}, status_code=500)

    kwargs = dict(
        start=str(start) if start else None,
        end=str(end) if end else None,
        vendor=vendor,
        category=category,
        min_amt=min_amount,
        max_amt=max_amount,
    )

    # write_only: satırlar disk'e akış halinde yazılır, tüm workbook RAM'de tutulmaz
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Faturalar")

    headers = ["ID", "Dosya", "Firma", "Tarih", "Saat", "Tutar", "KDV%",
               "KDV Tutarı", "Fatura No", "Kategori", "Ödeme", "QR", "Timestamp"]

    hdr_fill = PatternFill("solid", fgColor="2563EB")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    hdr_cells = []
    for h in headers:
        cell = openpyxl.cell.WriteOnlyCell(ws, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        hdr_cells.append(cell)
    ws.append(hdr_cells)

    written = 0
    truncated = False
    for row in iter_rows(**kwargs):
        if written >= _EXCEL_ROW_LIMIT:
            truncated = True
            break
        ws.append([
            row["id"] or "", row["filename"] or "",
            row["vendor"] or "", row["date"] or "", row["time"] or "",
            row["total"] or 0, row["vat_rate"] or "",
            row["vat_amount"] or 0, row["invoice_number"] or "",
            row["category"] or "", row["payment_method"] or "",
            (row["qr_raw"] or "")[:200], row["timestamp"] or "",
        ])
        written += 1

    if truncated:
        note = openpyxl.cell.WriteOnlyCell(ws, value=(
            f"⚠ Excel limiti: ilk {_EXCEL_ROW_LIMIT:,} satır gösterildi. "
            f"Tüm veri için /stats/export/csv kullanın."
        ))
        note.font = Font(bold=True, color="FF0000")
        ws.append([note])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="autotax_faturalar.xlsx"'},
    )


# ─── GET /stats/export/csv  (true streaming, sınır yok) ───
@router.get("/export/csv")
def export_csv(
    start:      Optional[date]  = Query(None),
    end:        Optional[date]  = Query(None),
    vendor:     Optional[str]   = Query(None, max_length=100),
    category:   Optional[str]   = Query(None, max_length=50),
    payment:    Optional[str]   = Query(None, max_length=50),
    min_amount: Optional[float] = Query(None, ge=0),
    max_amount: Optional[float] = Query(None, ge=0),
):
    """HTTP chunked streaming CSV — 100M satırda RAM kullanımı sabit (~2 MB)."""

    kwargs = dict(
        start=str(start) if start else None,
        end=str(end) if end else None,
        vendor=vendor,
        category=category,
        payment=payment,
        min_amt=min_amount,
        max_amt=max_amount,
    )

    def _gen():
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["ID", "Dosya", "Firma", "Tarih", "Saat", "Tutar",
                         "KDV%", "KDV_Tutar", "Fatura_No", "Kategori",
                         "Odeme", "QR", "Timestamp"])
        yield buf.getvalue()

        for row in iter_rows(**kwargs):
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow([
                row["id"] or "", row["filename"] or "",
                row["vendor"] or "", row["date"] or "", row["time"] or "",
                row["total"] or 0, row["vat_rate"] or "",
                row["vat_amount"] or 0, row["invoice_number"] or "",
                row["category"] or "", row["payment_method"] or "",
                (row["qr_raw"] or "")[:200], row["timestamp"] or "",
            ])
            yield buf.getvalue()

    return StreamingResponse(
        _gen(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="autotax_faturalar.csv"',
            "X-Content-Type-Options": "nosniff",
        },
    )


# ─── GET /stats/export/review-queue-excel  (sadece inceleme bekleyenler) ───
@router.get("/export/review-queue-excel")
def export_review_queue_excel():
    """needs_review=1 olan faturaları Excel olarak indir."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "openpyxl kurulu değil"}, status_code=500)

    from datetime import date as _date
    today = _date.today().isoformat()

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("İnceleme Bekleyenler")

    headers = ["ID", "Dosya", "Firma", "Tarih", "Saat", "Tutar", "KDV%",
               "KDV Tutarı", "Fatura No", "Kategori", "Ödeme", "İnceleme Nedeni", "Timestamp"]

    hdr_fill = PatternFill("solid", fgColor="DC2626")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    hdr_cells = []
    for h in headers:
        cell = openpyxl.cell.WriteOnlyCell(ws, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
        hdr_cells.append(cell)
    ws.append(hdr_cells)

    page, per_page = 1, 1000
    written = 0
    while True:
        result = get_review_queue(page=page, per_page=per_page)
        rows = result.get("invoices", [])
        if not rows:
            break
        for row in rows:
            ws.append([
                row.get("id") or "", row.get("filename") or "",
                row.get("vendor") or "", row.get("date") or "", row.get("time") or "",
                row.get("total") or "", row.get("vat_rate") or "",
                row.get("vat_amount") or "", row.get("invoice_number") or "",
                row.get("category") or "", row.get("payment_method") or "",
                row.get("review_reason") or "", row.get("timestamp") or "",
            ])
            written += 1
        if page >= result.get("pages", 1):
            break
        page += 1

    if written == 0:
        summary = openpyxl.cell.WriteOnlyCell(ws, value="İnceleme bekleyen fatura yok.")
        summary.font = Font(bold=True, color="16A34A")
        ws.append([summary])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"autotax_inceleme_bekleyen_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── GET /stats/export/review-queue-csv  (streaming, sınırsız) ───
@router.get("/export/review-queue-csv")
def export_review_queue_csv():
    """needs_review=1 olan tüm faturaları CSV olarak akış halinde indir."""
    from datetime import date as _date
    today = _date.today().isoformat()

    def _gen():
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["ID", "Dosya", "Firma", "Tarih", "Saat", "Tutar",
                         "KDV%", "KDV_Tutar", "Fatura_No", "Kategori",
                         "Odeme", "Inceleme_Nedeni", "Timestamp"])
        yield buf.getvalue()

        page, per_page = 1, 2000
        while True:
            result = get_review_queue(page=page, per_page=per_page)
            rows = result.get("invoices", [])
            if not rows:
                break
            for row in rows:
                buf2 = io.StringIO()
                w2 = csv.writer(buf2)
                w2.writerow([
                    row.get("id") or "", row.get("filename") or "",
                    row.get("vendor") or "", row.get("date") or "", row.get("time") or "",
                    row.get("total") or "", row.get("vat_rate") or "",
                    row.get("vat_amount") or "", row.get("invoice_number") or "",
                    row.get("category") or "", row.get("payment_method") or "",
                    row.get("review_reason") or "", row.get("timestamp") or "",
                ])
                yield buf2.getvalue()
            if page >= result.get("pages", 1):
                break
            page += 1

    fname = f"autotax_inceleme_bekleyen_{today}.csv"
    return StreamingResponse(
        _gen(),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Content-Type-Options": "nosniff",
        },
    )


# ─── GET /stats/ledger  (Muhasebe Defteri) ─────────────────────────────────
@router.get("/ledger")
def ledger(
    start:  Optional[str] = Query(None),
    end:    Optional[str] = Query(None),
    vendor: Optional[str] = Query(None),
    page:   int = Query(1,  ge=1),
    per_page: int = Query(50, ge=1, le=500),
):
    """Gelir + Gider muhasebe defteri. NET bakiyeyi de hesaplar."""
    import sqlite3
    from app.services.invoice_db import DB_PATH

    conditions = []
    params: list = []

    if start:
        conditions.append("date >= ?"); params.append(start)
    if end:
        conditions.append("date <= ?"); params.append(end)
    if vendor:
        conditions.append("vendor LIKE ?"); params.append(f"%{vendor}%")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    with sqlite3.connect(str(DB_PATH)) as con:
        con.row_factory = sqlite3.Row

        # Toplam sayı
        total = con.execute(
            f"SELECT COUNT(*) FROM invoices {where}", params
        ).fetchone()[0]

        # Gelir / Gider özeti
        agg = con.execute(f"""
            SELECT
                COALESCE(SUM(CASE WHEN invoice_type='income'  THEN total ELSE 0 END), 0) AS total_income,
                COALESCE(SUM(CASE WHEN invoice_type='expense' THEN total ELSE 0 END), 0) AS total_expense,
                COALESCE(SUM(CASE WHEN invoice_type='income'  THEN vat_amount ELSE 0 END), 0) AS vat_income,
                COALESCE(SUM(CASE WHEN invoice_type='expense' THEN vat_amount ELSE 0 END), 0) AS vat_expense,
                COUNT(CASE WHEN invoice_type='income'  THEN 1 END) AS count_income,
                COUNT(CASE WHEN invoice_type='expense' THEN 1 END) AS count_expense
            FROM invoices {where}
        """, params).fetchone()

        # Aylık özet
        monthly = con.execute(f"""
            SELECT
                SUBSTR(date,1,7) AS month,
                COALESCE(SUM(CASE WHEN invoice_type='income'  THEN total ELSE 0 END),0) AS income,
                COALESCE(SUM(CASE WHEN invoice_type='expense' THEN total ELSE 0 END),0) AS expense,
                COUNT(*) AS count
            FROM invoices {where}
            GROUP BY SUBSTR(date,1,7)
            ORDER BY month DESC
        """, params).fetchall()

        # Sayfalı fatura listesi
        offset = (page - 1) * per_page
        rows = con.execute(
            f"""SELECT id, filename, vendor, date, time, total, vat_amount,
                       invoice_number, category, payment_method, invoice_type,
                       needs_review, timestamp
                FROM invoices {where}
                ORDER BY date DESC, timestamp DESC
                LIMIT ? OFFSET ?""",
            params + [per_page, offset]
        ).fetchall()

    total_income  = round(float(agg["total_income"]),  2)
    total_expense = round(float(agg["total_expense"]), 2)
    net           = round(total_income - total_expense, 2)

    return {
        "count":         total,
        "page":          page,
        "per_page":      per_page,
        "pages":         max(1, -(-total // per_page)),
        "total_income":  total_income,
        "total_expense": total_expense,
        "vat_income":    round(float(agg["vat_income"]),  2),
        "vat_expense":   round(float(agg["vat_expense"]), 2),
        "count_income":  agg["count_income"],
        "count_expense": agg["count_expense"],
        "net":           net,
        "net_label":     "KAR" if net >= 0 else "ZARAR",
        "monthly": [
            {
                "month":   r["month"],
                "income":  round(float(r["income"]),  2),
                "expense": round(float(r["expense"]), 2),
                "net":     round(float(r["income"]) - float(r["expense"]), 2),
                "count":   r["count"],
            }
            for r in monthly
        ],
        "invoices": [
            {
                "id":             r["id"],
                "vendor":         r["vendor"] or "",
                "date":           r["date"] or "",
                "time":           r["time"] or "",
                "total":          r["total"],
                "vat_amount":     r["vat_amount"],
                "invoice_number": r["invoice_number"] or "",
                "category":       r["category"] or "",
                "payment_method": r["payment_method"] or "",
                "invoice_type":   r["invoice_type"] or "expense",
                "needs_review":   bool(r["needs_review"]),
                "filename":       r["filename"] or "",
            }
            for r in rows
        ],
    }


# ─── GET /stats/export/ledger-excel  (Muhasebe defteri Excel) ───────────────
@router.get("/export/ledger-excel")
def export_ledger_excel(
    start:  Optional[str] = Query(None),
    end:    Optional[str] = Query(None),
    vendor: Optional[str] = Query(None),
):
    """Tüm gelir+gider faturalarını muhasebe formatında Excel'e aktar."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "openpyxl kurulu değil"}, status_code=500)

    from datetime import date as _date
    today = _date.today().isoformat()

    data = ledger(start=start, end=end, vendor=vendor, page=1, per_page=1)

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Muhasebe Defteri")

    # Özet satırları
    green  = PatternFill("solid", fgColor="16A34A")
    red    = PatternFill("solid", fgColor="DC2626")
    blue   = PatternFill("solid", fgColor="1D4ED8")
    white  = Font(color="FFFFFF", bold=True, size=12)

    def _cell(ws, val, fill=None, font=None):
        cell = openpyxl.cell.WriteOnlyCell(ws, value=val)
        if fill: cell.fill = fill
        if font: cell.font = font
        return cell

    ws.append([_cell(ws, "MUHASEBE DEFTERİ", blue, white)])
    ws.append([_cell(ws, f"Rapor tarihi: {today}")])
    ws.append([])
    ws.append([
        _cell(ws, f"Toplam Gelir: {data['total_income']}", green, Font(color="FFFFFF", bold=True)),
        _cell(ws, f"Toplam Gider: {data['total_expense']}", red,   Font(color="FFFFFF", bold=True)),
        _cell(ws, f"NET {data['net_label']}: {data['net']}",
              PatternFill("solid", fgColor="16A34A" if data["net"] >= 0 else "DC2626"),
              Font(color="FFFFFF", bold=True)),
    ])
    ws.append([])

    # Aylık özet
    ws.append([_cell(ws, "AYLIK ÖZET", blue, white)])
    ws.append(["Ay", "Gelir", "Gider", "NET", "Fatura Sayısı"])
    for m in data["monthly"]:
        ws.append([m["month"], m["income"], m["expense"], m["net"], m["count"]])
    ws.append([])

    # Fatura listesi başlığı
    ws.append([_cell(ws, "FATURA LİSTESİ", blue, white)])
    hdr_fill = PatternFill("solid", fgColor="374151")
    hdr_font = Font(color="FFFFFF", bold=True)
    ws.append([
        _cell(ws, h, hdr_fill, hdr_font)
        for h in ["Tür", "Tarih", "Firma", "Tutar", "KDV", "Kategori", "Ödeme", "Fatura No"]
    ])

    # Tüm sayfalarda fatura yaz
    import sqlite3
    from app.services.invoice_db import DB_PATH
    conditions, params = [], []
    if start:  conditions.append("date >= ?"); params.append(start)
    if end:    conditions.append("date <= ?"); params.append(end)
    if vendor: conditions.append("vendor LIKE ?"); params.append(f"%{vendor}%")
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    with sqlite3.connect(str(DB_PATH)) as con:
        con.row_factory = sqlite3.Row
        rows = con.execute(
            f"SELECT * FROM invoices {where} ORDER BY date DESC", params
        ).fetchall()

    for r in rows:
        typ  = "GELİR" if r["invoice_type"] == "income" else "GİDER"
        fill = PatternFill("solid", fgColor="DCFCE7") if r["invoice_type"] == "income" else PatternFill("solid", fgColor="FEE2E2")
        ws.append([
            _cell(ws, typ,                     fill),
            _cell(ws, r["date"] or "",         fill),
            _cell(ws, r["vendor"] or "",       fill),
            _cell(ws, r["total"],              fill),
            _cell(ws, r["vat_amount"],         fill),
            _cell(ws, r["category"] or "",     fill),
            _cell(ws, r["payment_method"] or "", fill),
            _cell(ws, r["invoice_number"] or "", fill),
        ])

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"autotax_muhasebe_{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
